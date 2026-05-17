import sys
sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.font_manager as fm
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import io, os

# ── 한글 폰트 설정 ─────────────────────────────────────
font_candidates = [
    'C:/Windows/Fonts/malgun.ttf',
    'C:/Windows/Fonts/NanumGothic.ttf',
]
for fp in font_candidates:
    if os.path.exists(fp):
        fm.fontManager.addfont(fp)
        prop = fm.FontProperties(fname=fp)
        plt.rcParams['font.family'] = prop.get_name()
        break
plt.rcParams['axes.unicode_minus'] = False

# ── 진도 단계 수치화 ───────────────────────────────────
LEVEL_MAP = {
    'A':1,'B':2,'C':3,'D':4,'E':5,'F':6,
    'G':7,'H':8,'I':9,'J':10,'K':11,'L':12,'M':13
}
LEVEL_LABEL = {
    1:'A(초1)',2:'B(초2)',3:'C(초3)',4:'D(초4)',5:'E(초5)',6:'F(초6)',
    7:'G(중1)',8:'H(중2)',9:'I(중3)',10:'J(고1)',11:'K(고2)',12:'L(고3)',13:'M(수능)'
}
GRADE_MAP = {'초1':1,'초2':2,'초3':3,'초4':4,'초5':5,'초6':6,
             '중1':7,'중2':8,'중3':9,'고1':10,'고2':11,'고3':12}

df = pd.read_excel('C:/Users/najun/Downloads/진도상위회원 조회__20260413.xlsx', header=1)
df.columns = ['소속','교사번호','교사명','회원코드','과목','회원명','학년',
              '현재진도','진도수준','학습개월','학습개월2','년월','단계','합격여부','기타']

df['학년_수치'] = df['학년'].map(GRADE_MAP)
df['진도_수치'] = df['현재진도'].map(LEVEL_MAP)
df['갭'] = df['진도_수치'] - df['학년_수치']
df['대회합격'] = df['합격여부'].fillna('미참').apply(lambda x: '합격' if x == '합격' else '미참')

# ══════════════════════════════════════════════════════
# 그래프 1: 학년 vs 진도 수준 갭 (막대 차트)
# ══════════════════════════════════════════════════════
fig1, ax1 = plt.subplots(figsize=(14, 7))
fig1.patch.set_facecolor('#F8F9FA')
ax1.set_facecolor('#F8F9FA')

names = df['회원명'].tolist()
grades = df['학년_수치'].tolist()
levels = df['진도_수치'].tolist()
gaps   = df['갭'].tolist()
colors = ['#C00000' if g >= 5 else '#ED7D31' if g >= 4 else '#2E75B6' for g in gaps]

x = np.arange(len(names))
bars_g = ax1.bar(x - 0.2, grades, 0.35, label='학년 수준', color='#BDD7EE', zorder=3)
bars_l = ax1.bar(x + 0.2, levels, 0.35, label='현재 진도', color='#2E75B6', zorder=3)

for i, (g, l, gap) in enumerate(zip(grades, levels, gaps)):
    ax1.annotate('', xy=(x[i]+0.2, l), xytext=(x[i]-0.2, g),
                 arrowprops=dict(arrowstyle='->', color='#C00000', lw=1.5))
    ax1.text(x[i]+0.55, (g+l)/2, f'+{gap}', color='#C00000',
             fontsize=9, fontweight='bold', va='center')

ax1.set_xticks(x)
ax1.set_xticklabels([f"{n}\n({df['학년'].iloc[i]})" for i, n in enumerate(names)],
                    fontsize=9)
ax1.set_yticks(range(1, 14))
ax1.set_yticklabels([LEVEL_LABEL[i] for i in range(1, 14)], fontsize=8)
ax1.set_title('관양3지구 진도 상위 회원 — 학년 대비 진도 수준', fontsize=15, fontweight='bold', pad=15)
ax1.set_ylabel('단계', fontsize=11)
ax1.legend(fontsize=10, loc='upper left')
ax1.grid(axis='y', alpha=0.3, zorder=0)
ax1.spines[['top','right']].set_visible(False)

p1 = mpatches.Patch(color='#C00000', label='갭 5단계 이상 (초긴급)')
p2 = mpatches.Patch(color='#ED7D31', label='갭 4단계 (긴급)')
p3 = mpatches.Patch(color='#2E75B6', label='갭 3단계 (우선)')
ax1.legend(handles=[p1, p2, p3], fontsize=9, loc='upper left')

plt.tight_layout()
graph1_path = 'd:/coding/kumon-20251010/graph_진도갭.png'
plt.savefig(graph1_path, dpi=150, bbox_inches='tight')
plt.close()

# ══════════════════════════════════════════════════════
# 그래프 2: 교사별 상위 회원 분포 파이차트
# ══════════════════════════════════════════════════════
fig2, (ax2a, ax2b) = plt.subplots(1, 2, figsize=(13, 6))
fig2.patch.set_facecolor('#F8F9FA')

teacher_cnt = df['교사명'].value_counts()
colors_pie = ['#2E75B6','#ED7D31','#70AD47','#C00000','#7030A0','#FFB900']
wedges, texts, autotexts = ax2a.pie(
    teacher_cnt.values, labels=teacher_cnt.index, autopct='%1.0f%%',
    colors=colors_pie[:len(teacher_cnt)], startangle=90,
    textprops={'fontsize': 11})
for at in autotexts:
    at.set_fontweight('bold')
ax2a.set_title('교사별 진도상위 회원 수', fontsize=13, fontweight='bold')

gap_group = ['5단계+' if g >= 5 else '4단계' if g == 4 else '3단계' for g in gaps]
gap_cnt = pd.Series(gap_group).value_counts()
order = ['5단계+','4단계','3단계']
gap_cnt = gap_cnt.reindex([o for o in order if o in gap_cnt.index])
bar_colors = ['#C00000','#ED7D31','#2E75B6'][:len(gap_cnt)]
ax2b.bar(gap_cnt.index, gap_cnt.values, color=bar_colors, width=0.5, zorder=3)
for i, v in enumerate(gap_cnt.values):
    ax2b.text(i, v + 0.1, f'{v}명', ha='center', fontsize=13, fontweight='bold')
ax2b.set_title('갭 규모별 분포', fontsize=13, fontweight='bold')
ax2b.set_ylabel('회원 수', fontsize=11)
ax2b.set_facecolor('#F8F9FA')
ax2b.grid(axis='y', alpha=0.3, zorder=0)
ax2b.spines[['top','right']].set_visible(False)

plt.tight_layout()
graph2_path = 'd:/coding/kumon-20251010/graph_분포.png'
plt.savefig(graph2_path, dpi=150, bbox_inches='tight')
plt.close()

# ══════════════════════════════════════════════════════
# 엑셀 보고서 생성
# ══════════════════════════════════════════════════════
wb = Workbook()

# 색상/스타일 헬퍼
DARK   = 'FF1F3564'
MID    = 'FF2E75B6'
RED    = 'FFC00000'
ORANGE = 'FFED7D31'
GREEN  = 'FF70AD47'
LGRAY  = 'FFF2F2F2'
WHITE  = 'FFFFFFFF'

def cell_style(ws, row, col, value='', bg=None, fg='FF000000',
               bold=False, size=11, align='center', border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(bold=bold, size=size, color=fg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if border:
        s = Side(style='thin', color='FFBFBFBF')
        c.border = Border(left=s, right=s, top=s, bottom=s)
    return c

# ── 시트 1: 진도갭 분석 ───────────────────────────────
ws1 = wb.active
ws1.title = '진도갭 분석'

ws1.merge_cells('A1:L1')
cell_style(ws1, 1, 1, '관양3지구 진도 상위 회원 — 학년 대비 진도 갭 분석', DARK, WHITE, True, 14)
ws1.row_dimensions[1].height = 32

headers = ['교사명','회원명','학년','과목','현재진도','진도수준','학년수준(숫자)',
           '진도수준(숫자)','갭(단계)','학습개월','최종대회','대회결과']
col_w   = [10,8,6,6,8,12,12,12,10,10,10,10]
for i, (h, w) in enumerate(zip(headers, col_w), 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
    cell_style(ws1, 2, i, h, MID, WHITE, True, 10)
ws1.row_dimensions[2].height = 22

for r, row in enumerate(df.itertuples(), 3):
    gap = row.갭
    bg = RED if gap >= 5 else ORANGE if gap == 4 else 'FF2E75B6' if gap == 3 else LGRAY
    fg = WHITE if gap >= 3 else 'FF000000'
    data = [row.교사명, row.회원명, row.학년, row.과목, row.현재진도,
            row.진도수준, row.학년_수치, row.진도_수치, gap,
            int(row.학습개월) if not pd.isna(row.학습개월) else '-',
            str(row.단계) if not pd.isna(row.단계) else '-',
            row.대회합격]
    for c, v in enumerate(data, 1):
        cell_style(ws1, r, c, v, bg if c == 9 else (LGRAY if r % 2 == 0 else WHITE),
                   fg if c == 9 else 'FF000000', c == 9, 10)
    ws1.row_dimensions[r].height = 20

# 그래프 삽입
img1 = XLImage(graph1_path)
img1.width, img1.height = 900, 420
ws1.add_image(img1, 'A24')

img2 = XLImage(graph2_path)
img2.width, img2.height = 820, 360
ws1.add_image(img2, 'A54')

# ── 시트 2: 마스타 입회 상담 전략 ─────────────────────
ws2 = wb.create_sheet('마스타 입회 상담 전략')
for i, w in enumerate([5,22,55,55], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells('A1:D1')
cell_style(ws2, 1, 1, '마스타 입회 상담 전략 — 진도 갭 활용법', DARK, WHITE, True, 14)
ws2.row_dimensions[1].height = 32

# 전략 데이터
strategies = [
    ('1단계', 'OPEN — 그래프 보여주기',
     '• "오늘 저희 지구 진도 상위 회원 현황을 보여드리고 싶어서요"\n'
     '• 진도 갭 그래프를 출력해서 어머니께 제시\n'
     '• "○○이가 현재 이 위치에 있습니다 (가리키며)"',
     '학부모의 호기심·자부심 유발\n그래프가 말하게 하고, 선생님은 듣기'),

    ('2단계', 'FACT — 갭 수치로 강점 각인',
     '• "학년 대비 ○단계 앞서 있어요. 이 말은 ○학년 수업이 이미 쉽다는 뜻입니다"\n'
     '• 갭이 클수록 강조: 5단계+ → "중학교/고등학교 수준을 지금 배우고 있어요"\n'
     '• 최종대회 합격 단계 언급: "이미 △단계 인정받았습니다"',
     '숫자로 말하기 — 추상적 칭찬 X\n"잘 하네요" 대신 "+4단계"로 구체화'),

    ('3단계', 'NEED — 마스타의 필요성 연결',
     '• "이 진도라면 일반 구몬 학습 속도가 느릴 수 있어요"\n'
     '• "마스타는 이런 상위 진도 회원을 위해 설계된 심화 과정입니다"\n'
     '• "같은 수준 친구들과 함께 배우면 동기부여가 훨씬 커집니다"',
     '"느리다"는 위기감 + "해결책"으로 마스타 포지셔닝\n불안 자극이 아닌 성장 기회로 프레이밍'),

    ('4단계', 'PROOF — 성공 사례 제시',
     '• 기존 합격 회원 사례 공유 (개인정보 주의)\n'
     '• "지구 내 비슷한 진도였던 ○○이가 마스타 후 ○○대 진학했습니다"\n'
     '• 대회 합격 단계를 활용: "이미 M단계 합격 = 수능 수준"',
     '구체적 증거 > 말뿐인 설명\n이미 합격한 회원 데이터가 최고의 설득 도구'),

    ('5단계', 'CLOSE — 입회 제안',
     '• "다음 달 마스타 입회 자리가 ○자리 남아 있어요"\n'
     '• "우선 체험 수업 한 번만 해보시겠어요?"\n'
     '• 거절 시: "그럼 현재 학습 목표를 한번 같이 세워볼까요?"',
     '선택지 제한으로 결정 장벽 낮추기\n체험 → 입회의 작은 발걸음 전략'),
]

row = 2
colors_s = [MID, GREEN, ORANGE, RED, 'FF7030A0']
for i, (step, title, script, tip) in enumerate(strategies):
    bg = colors_s[i]
    # 단계 헤더
    ws2.merge_cells(f'A{row}:D{row}')
    cell_style(ws2, row, 1, f'{step}  |  {title}', bg, WHITE, True, 12)
    ws2.row_dimensions[row].height = 28
    row += 1

    # 상담 스크립트 / 포인트
    cell_style(ws2, row, 2, '상담 스크립트', bg, WHITE, True, 10, 'center')
    cell_style(ws2, row, 3, '핵심 포인트', bg, WHITE, True, 10, 'center')
    ws2.merge_cells(f'A{row}:A{row+4}')
    ws2.row_dimensions[row].height = 18
    row += 1

    for line in script.strip().split('\n'):
        cell_style(ws2, row, 2, line.strip(), LGRAY if row%2==0 else WHITE,
                   'FF000000', False, 10, 'left', wrap=True)
        ws2.row_dimensions[row].height = 22
        row += 1

    cell_style(ws2, row-len(script.split('\n')), 3, tip,
               'FFFFF2CC', 'FF000000', False, 10, 'left', wrap=True)
    ws2.merge_cells(f'C{row-len(script.split(chr(10)))}:C{row-1}')

    row += 1

# ── 시트 3: 개인별 상담 카드 ──────────────────────────
ws3 = wb.create_sheet('개인별 상담 카드')
ws3.merge_cells('A1:G1')
cell_style(ws3, 1, 1, '개인별 마스타 입회 상담 포인트', DARK, WHITE, True, 14)
ws3.row_dimensions[1].height = 32

for i, w in enumerate([10,8,6,6,10,12,50], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

headers3 = ['교사명','회원명','학년','진도','갭','대회','상담 포인트']
for i, h in enumerate(headers3, 1):
    cell_style(ws3, 2, i, h, MID, WHITE, True, 10)
ws3.row_dimensions[2].height = 22

def make_talk(row):
    name = row['회원명']
    grade = row['학년']
    level = row['현재진도']
    gap = row['갭']
    jindo = row['진도수준']
    month = row['학습개월']
    contest = row['단계'] if not pd.isna(row['단계']) else None

    talk = f'"{name} 회원은 {grade}인데 현재 {jindo}({level}단계) 학습 중입니다.'
    talk += f' 학년보다 {gap}단계 앞서 있어요.'
    if contest:
        talk += f' 최종대회 {contest}단계 합격 이력도 있습니다.'
    if gap >= 5:
        talk += f' 이미 고등 수준 학습 중 — 마스타 전환 강력 추천."'
    elif gap >= 4:
        talk += f' 중학교 수준 완전 선행 — 마스타 입회 적기입니다."'
    else:
        talk += f' 꾸준한 선행 진도 — 마스타로 심화 도전 가능합니다."'
    return talk

for r, (_, row) in enumerate(df.iterrows(), 3):
    gap = row['갭']
    bg = 'FFFFC7CE' if gap >= 5 else 'FFFFEB9C' if gap == 4 else 'FFC6EFCE'
    data = [row['교사명'], row['회원명'], row['학년'], row['현재진도'],
            f'+{gap}단계',
            str(row['단계']) + '합격' if not pd.isna(row['단계']) else '미참',
            make_talk(row)]
    for c, v in enumerate(data, 1):
        align = 'left' if c == 7 else 'center'
        cell_style(ws3, r, c, v, bg, 'FF000000', False, 10, align, wrap=(c==7))
    ws3.row_dimensions[r].height = 48

out = 'C:/Users/najun/Downloads/마스타_입회_상담_전략.xlsx'
wb.save(out)
print(f'완료: {out}')
print(f'파일 크기: {os.path.getsize(out):,} bytes')
